import time
import logging
import sys
from engine_v2 import run_engine, get_sheet, read_pending_leads, generate_batch
from mailer_v2 import send_email

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S"
)
logger = logging.getLogger(__name__)


def main_menu():
    while True:
        try:
            import config
            sheet = get_sheet()
            leads_all = read_pending_leads(sheet, n=500)
            leads_listos = [l for l in leads_all if l.get(config.COL_ESTADO) == "LISTO"]

            print("\n================================================")
            print("🎯 ANTIGRAVITY MAILER — PANEL DE CONTROL")
            print("================================================")
            print(f"📋 Leads listos: {len(leads_listos)}")
            print("================================================")
            print("  [1] Envío de prueba (1 lead específico)")
            print("  [2] Ver leads disponibles")
            print("  [3] Generar Excel de auditoría (50 correos)")
            print("  [4] Enviar lote aprobado")
            print("  [5] Envío de prueba random a mi email")
            print("  [6] AUTO-TEST END-TO-END (Envío rápido + simular click)")
            print("  [0] Salir")
            print("================================================")

            choice = input("Tu elección: ").strip()

            if choice == "1":
                print("\n------------------------------------------------")
                print("LEADS DISPONIBLES PARA PRUEBA:")
                # Priorizamos Fila 44 para la prueba solicitada por el usuario
                f44 = [l for l in leads_all if str(l.get("_fila")) == "44"]
                others = [l for l in leads_all if str(l.get("_fila")) != "44"]
                current_leads = (f44 + others)[:10]
                for i, l in enumerate(current_leads, 1):
                    empresa = l.get("EMPRESA", "Sin nombre")
                    fila = l.get("_fila", "?")
                    print(f"[{i}] {empresa} (Fila {fila})")

                try:
                    target_idx = int(input("\n¿Qué lead quieres probar? (número): ")) - 1
                    if 0 <= target_idx < len(current_leads):
                        lead = current_leads[target_idx]
                        logger.info(f"🤖 Generando previsualización para {lead.get('EMPRESA')}...")
                        results = generate_batch([lead])
                        if results and not results[0].get("ai_error"):
                            ai = results[0]
                            print("\n------------------------------------------------")
                            print("📧 CORREO GENERADO:")
                            print(f"SUBJECT:  {ai.get('Asunto_Email', ai.get('_raw_SUBJECT', ''))}")
                            print(f"HOOK:     {ai.get('Intro_Equipo', ai.get('_raw_HOOK', ''))}")
                            print(f"COMP:     {ai.get('Parrafo_Mercado', ai.get('_raw_COMPETIDOR', ''))}")
                            print(f"GAP:      {ai.get('_raw_GAP', ai.get('Parrafo_Emocional', ''))}")
                            print(f"INJURY:   {ai.get('_raw_INJURY', '')}")
                            print(f"QUESTION: {ai.get('Pregunta_Cierre', ai.get('_raw_QUESTION', ''))}")
                            print(f"PD:       {ai.get('PD_Urgencia', ai.get('_raw_PD', ''))}")
                            print("------------------------------------------------")

                            confirm = input(f"¿Enviar este correo a {lead.get('EMAIL')}? (s/n): ").strip().lower()
                            if confirm == "s":
                                if send_email(lead, ai):
                                    logger.info("✅ Email de prueba enviado.")
                                else:
                                    logger.error("❌ Falló el envío del email.")
                        else:
                            logger.error("❌ Falló la generación de contenido por IA.")
                    else:
                        print("Número fuera de rango.")
                except ValueError:
                    print("Entrada no válida.")

            elif choice == "2":
                print("\n------------------------------------------------")
                print(f"{'EMPRESA':<30} | {'NICHO':<15} | {'CIUDAD':<15} | {'PRIORIDAD'}")
                print("-" * 80)
                for l in leads_all:
                    empresa = str(l.get("EMPRESA", ""))[:30]
                    nicho = str(l.get("NICHO", ""))[:15]
                    ciudad = str(l.get("CIUDAD", l.get("ESTADO_GEO", "N/D")))[:15]
                    prioridad = l.get("PRIORIDAD", "0")
                    print(f"{empresa:<30} | {nicho:<15} | {ciudad:<15} | {prioridad}")
                input("\nPresiona Enter para volver al menú...")

            elif choice == "3":
                print("\n------------------------------------------------")
                print("🔄 GENERADOR DE EXCEL DE AUDITORÍA")
                print("------------------------------------------------")
                print("¿Cuántos correos deseas generar para auditar?")
                print("(Escribe un número, ej: 10, 20, 50, o 'todo')")
                
                cantidad_input = input("Cantidad a auditar: ").strip().lower()
                
                if cantidad_input == 'todo':
                    cantidad = 500
                else:
                    try:
                        cantidad = int(cantidad_input)
                        if cantidad <= 0:
                            print("❌ La cantidad debe ser mayor a 0.")
                            continue
                    except ValueError:
                        print("❌ Entrada no válida.")
                        continue

                sheet = get_sheet()
                leads = read_pending_leads(sheet, n=cantidad)

                if not leads:
                    print("❌ No hay leads pendientes.")
                    continue

                print(f"⚙️  Generando {len(leads)} correos — espera...")
                results = generate_batch(leads)

                import openpyxl
                from openpyxl.styles import Font, PatternFill
                from datetime import datetime

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Auditoría"

                headers = [
                    "FILA", "EMPRESA", "EMAIL", "NICHO", "CIUDAD",
                    "SUBJECT", "HOOK", "COMPETITOR", "GAP",
                    "INJURY", "QUESTION", "PD", "IDIOMA_OK"
                ]
                ws.append(headers)

                for col in range(1, len(headers) + 1):
                    ws.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF")
                    ws.cell(row=1, column=col).fill = PatternFill("solid", fgColor="1a1a2e")

                palabras_esp = [
                    "liberando", "tiempo", "negocio", "construyendo",
                    "aquí", "también", "cómo", "están", "también"
                ]

                for i, r in enumerate(results, 2):
                    subject  = r.get("Asunto_Email", "")
                    hook     = r.get("Intro_Equipo", "")
                    comp     = r.get("Parrafo_Mercado", "")
                    gap      = r.get("Parrafo_Emocional", "")
                    injury   = r.get("_raw_INJURY", "")
                    question = r.get("Pregunta_Cierre", "")
                    pd_      = r.get("PD_Urgencia", "")

                    texto = (subject + hook + comp + gap + injury + question + pd_).lower()
                    idioma_ok = "✅ EN" if not any(
                        p in texto for p in palabras_esp
                    ) else "⚠️ ES"

                    ws.append([
                        r.get("_fila", ""),
                        r.get("EMPRESA", ""),
                        r.get("EMAIL", ""),
                        r.get("NICHO", ""),
                        r.get("CIUDAD", ""),
                        subject, hook, comp, gap,
                        injury, question, pd_, idioma_ok
                    ])

                    if "⚠️" in idioma_ok:
                        for col in range(1, len(headers) + 1):
                            ws.cell(row=i, column=col).fill = PatternFill(
                                "solid", fgColor="FFE0E0"
                            )

                for col_letter, width in [
                    ("A", 8), ("B", 30), ("C", 30), ("D", 20), ("E", 15),
                    ("F", 50), ("G", 60), ("H", 60), ("I", 50),
                    ("J", 50), ("K", 50), ("L", 50), ("M", 12)
                ]:
                    ws.column_dimensions[col_letter].width = width

                fecha = datetime.now().strftime("%Y%m%d_%H%M")
                path = (
                    f"C:/Users/GUEST/.gemini/antigravity/"
                    f"scratch/micro-saas-2/auditoria_{fecha}.xlsx"
                )
                wb.save(path)
                print(f"\n✅ Excel generado: auditoria_{fecha}.xlsx")
                print(f"📊 {len(results)} correos listos para auditar")
                print(f"📁 {path}")
                input("\nRevisa el Excel y presiona Enter para volver...")

            elif choice == "4":
                print("\n------------------------------------------------")
                print("🚀 ENVÍO DE LOTE (WARM-UP ESTRATÉGICO)")
                print("------------------------------------------------")
                print("¿Cuántos correos deseas enviar en esta tanda?")
                print("(Escribe un número, ej: 10, 20, 50, o 'todo')")
                
                cantidad_input = input("Cantidad a enviar: ").strip().lower()
                
                if cantidad_input == 'todo':
                    cantidad = 500
                else:
                    try:
                        cantidad = int(cantidad_input)
                        if cantidad <= 0:
                            print("❌ La cantidad debe ser mayor a 0.")
                            continue
                    except ValueError:
                        print("❌ Entrada no válida.")
                        continue

                logger.info(f"🚀 Iniciando envío manual de {cantidad} leads...")
                run_engine(n=cantidad)
                logger.info("✅ Lote procesado.")

            elif choice == "5":
                print("\n------------------------------------------------")
                print("📧 ENVÍO DE PRUEBA — CORREO RANDOM")
                print("------------------------------------------------")

                email_destino = input(
                    "¿A qué email enviar la prueba? "
                ).strip()

                if not email_destino or "@" not in email_destino:
                    print("❌ Email inválido.")
                    continue

                # Seleccionar lead random
                import random
                sheet = get_sheet()
                leads = read_pending_leads(sheet, n=50)

                if not leads:
                    print("❌ No hay leads disponibles.")
                    continue

                lead_random = random.choice(leads)
                empresa = lead_random.get("EMPRESA", "Sin nombre")
                print(f"\n🎲 Lead seleccionado al azar: {empresa}")
                print("⚙️  Generando correo...")

                results = generate_batch([lead_random])

                if not results or results[0].get("ai_error"):
                    print("❌ Error generando el correo.")
                    continue

                ai = results[0]
                print("\n------------------------------------------------")
                print("📧 CORREO GENERADO:")
                print(f"SUBJECT:  {ai.get('Asunto_Email', '')}")
                print(f"HOOK:     {ai.get('Intro_Equipo', '')}")
                print(f"COMP:     {ai.get('Parrafo_Mercado', '')}")
                print(f"GAP:      {ai.get('Parrafo_Emocional', '')}")
                print(f"INJURY:   {ai.get('_raw_INJURY', '')}")
                print(f"QUESTION: {ai.get('Pregunta_Cierre', '')}")
                print(f"PD:       {ai.get('PD_Urgencia', '')}")
                print("------------------------------------------------")

                confirm = input(
                    f"\n¿Enviar este correo a {email_destino}? (s/n): "
                ).strip().lower()

                if confirm == "s":
                    # Sobrescribir email del lead con el destino de prueba
                    lead_random["EMAIL"] = email_destino
                    if send_email(lead_random, ai):
                        print(f"✅ Correo de prueba enviado a {email_destino}")
                    else:
                        print("❌ Error en el envío.")

            elif choice == "6":
                print("\n------------------------------------------------")
                print("⚡ AUTO-TEST END-TO-END")
                print("------------------------------------------------")
                import os
                import random
                import requests
                import urllib.parse
                import config

                email_destino = os.getenv("TEST_EMAIL", "").strip()
                if not email_destino:
                    email_destino = input("TEST_EMAIL no definido en .env. Introduce tu email: ").strip()
                
                sheet = get_sheet()
                leads = read_pending_leads(sheet, n=50)

                if not leads:
                    print("❌ No hay leads pendientes.")
                    continue

                lead_random = random.choice(leads)
                empresa = lead_random.get("EMPRESA", "Sin nombre")
                print(f"\n🎲 Lead seleccionado al azar: {empresa}")
                print(f"📧 Reemplazando destino con: {email_destino}")
                
                # Sobrescribir el email
                lead_random["EMAIL"] = email_destino
                
                print("⚙️  Generando correo...")
                results = generate_batch([lead_random])

                if not results or results[0].get("ai_error"):
                    print("❌ Error generando el correo.")
                    continue

                ai = results[0]
                print(f"🚀 Enviando Correo 1 (Apertura) a {email_destino}...")
                
                if send_email(lead_random, ai):
                    print("✅ Correo 1 enviado con éxito.")
                    
                    # Construir URL de tracking para simular click
                    nombre_corto = " ".join(empresa.split()[:2])
                    kpi4_val = lead_random.get("KPI_EFFICIENCY_GAP", "")
                    kpi5_val = lead_random.get("KPI_RIVALRY", "")
                    comp_val = lead_random.get("COMPETIDOR_NAME") or lead_random.get("COMPETITOR_NAME") or ""
                    
                    params = urllib.parse.urlencode({
                        "email": email_destino,
                        "name":  nombre_corto,
                        "kpi4":  kpi4_val[:200] if kpi4_val else "",
                        "kpi5":  kpi5_val[:200] if kpi5_val else "",
                        "comp":  comp_val[:80]  if comp_val  else "",
                    })
                    track_url = f"{config.SERVER_BASE_URL}/track?{params}"
                    
                    print(f"\n🔗 Link de Tracking generado internamente:\n{track_url}")
                    
                    simular = input("\n¿Deseas simular el clic por consola AHORA (s) o prefieres ir a tu bandeja a leerlo y hacer clic tú mismo (n)? (s/n): ").strip().lower()
                    if simular == "s":
                        print(f"🌐 Haciendo GET a la nube de PythonAnywhere...")
                        try:
                            resp = requests.get(track_url, timeout=15)
                            print(f"✅ Respuesta del servidor: HTTP {resp.status_code}")
                            if resp.status_code == 200:
                                print("🎉 ¡El click fue registrado exitosamente!")
                                print("📩 Revisa tu bandeja de entrada en 10-20 segundos para ver el Correo 2 (El Regalo).")
                            else:
                                print("⚠️ El servidor no respondió con 200 OK.")
                        except Exception as e:
                            print(f"❌ Error simulando click: {e}")
                else:
                    print("❌ Error en el envío del Correo 1.")

            elif choice == "0":
                print("Saliendo...")
                sys.exit(0)
            else:
                print("Opción no válida.")

        except Exception as e:
            logger.error(f"Error en el menú: {e}")
            time.sleep(2)


if __name__ == "__main__":
    main_menu()
